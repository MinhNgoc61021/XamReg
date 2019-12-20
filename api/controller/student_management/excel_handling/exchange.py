from flask import (
    Blueprint,
    # Blueprint is a way to organize a group of related views and other code
    # There will be 2 blueprints: one for authentication and one for posts function
    request,
    jsonify
)
from db.entity_db import User, Subject, Student_Status, Log
from openpyxl import load_workbook
from controller.authentication.auth import token_required
from controller.time_conversion.asia_timezone import set_custom_log_time
import re

# excel handling for admin
excel_handling = Blueprint('import_export', __name__, url_prefix='/handling')


# to get excel data
@excel_handling.route('/upload', methods=['POST'])
@token_required
def upload(current_user):
    try:
        print(request.files['student_list_excel'], flush=True)
        # load file
        excel_file = load_workbook(request.files['student_list_excel'])
        # select Danh sách sinh viên.xlsx
        sheet = excel_file.active

        # get max row count
        max_row = sheet.max_row
        # get max column count
        max_column = sheet.max_column

        print(max_row, flush=True)
        print(max_column, flush=True)
        # When the excel file are included with student, subject info & status
        if max_column == 8:
            # set excel_data to get data in order to create new account, subject, qualification for students to SQLAlchemy
            for i in range(2, max_row + 1):
                excel_data = {
                    'ID': None,
                    'fullname': None,
                    'dob': None,
                    'gender': None,
                    'courseID': None,
                    'subjectID': None,
                    'subjectTitle': None,
                    'status': None,
                }
                # Iterate over the dict and all the columns
                for j, index in zip(range(1, max_column + 1), excel_data):
                    # add data to excel_data dict
                    excel_data[index] = sheet.cell(row=i, column=j).value
                print(excel_data, flush=True)

                # add students to database
                # check validation
                ID = re.search('^\d{8}$', str(excel_data['ID']).replace(' ', ''))
                fullname = re.search("^[a-zA-Z_ÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠàáâãèéêìíòóôõùúăđĩũơƯĂẠẢẤẦẨẪẬẮẰẲẴẶ" +
                                     "ẸẺẼỀẾỂưăạảấầẩẫậắằẳẵặẹẻẽềếểỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪễệỉịọỏốồổỗộớờởỡợ" +
                                     "ụủứừỬỮỰỲỴÝỶỸửữựỳýỵỷỹ\\s ]+$", str(excel_data['fullname']))
                dob = re.search('^(((0)[1-9])|((1)[0-2]))(\/)([0-2][0-9]|(3)[0-1])(\/)\d{4}',
                                str(excel_data['dob'].strftime('%m/%d/%Y, %H:%M:%S')))
                gender = re.search('(Nam|Nữ)', str(excel_data['gender']))
                courseID = re.search('^[K|k][1-9][0-9][A-Za-z]+[1-9]*', str(excel_data['courseID']))
                subjectID = re.search('(^(([A-Z]|[a-z]){3})([1-9][(0-9)]{3}))',
                                      str(excel_data['subjectID']))
                status = re.search('(đủ điều kiện|không đủ điều kiện)', excel_data['status'].lower())
                # print(ID, flush=True)
                # print(fullname, flush=True)
                # print(dob, flush=True)
                # print(gender, flush=True)
                # print(courseID, flush=True)
                # print(subjectID, flush=True)
                # print(status, flush=True)

                if (ID is not None) and (fullname is not None) and (dob is not None) and (gender is not None) and (
                        courseID is not None) and (subjectID is not None) and (status is not None):

                    User.create(str(excel_data['ID']).replace(' ', ''),
                                str(excel_data['ID']).replace(' ', '') + '@vnu.edu.vn',
                                str(excel_data['ID']).replace(' ', ''),
                                str(excel_data['fullname']).title(),
                                excel_data['dob'],
                                excel_data['gender'],
                                excel_data['courseID'],
                                'Student')
                    # add subject to database
                    Subject.create(excel_data['subjectID'],
                                   excel_data['subjectTitle'])
                    # add qualified and unqualified students to database
                    Student_Status.create(excel_data['ID'],
                                          excel_data['subjectID'], excel_data['status'])
                else:
                    return jsonify({'status': 'error', 'message': 'Du'}), 400

        # When the excel file are included of only student info
        elif max_column == 5:
            # set excel_data to get data in order to create new account only for students to SQLAlchemy
            for i in range(2, max_row + 1):
                excel_data = {
                    'ID': None,
                    'fullname': None,
                    'dob': None,
                    'gender': None,
                    'courseID': None,
                }
                # Iterate over the dict and all the columns
                for j, index in zip(range(1, max_column + 1), excel_data):
                    # add data to excel_data dict
                    excel_data[index] = sheet.cell(row=i, column=j).value
                print(excel_data, flush=True)

                # add students to database
                # check validation
                ID = re.search('^\d{8}$', str(excel_data['ID']).replace(' ', ''))
                fullname = re.search("^[a-zA-Z_ÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠàáâãèéêìíòóôõùúăđĩũơƯĂẠẢẤẦẨẪẬẮẰẲẴẶ" +
                                     "ẸẺẼỀẾỂưăạảấầẩẫậắằẳẵặẹẻẽềếểỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪễệỉịọỏốồổỗộớờởỡợ" +
                                     "ụủứừỬỮỰỲỴÝỶỸửữựỳýỵỷỹ\\s ]+$", str(excel_data['fullname']))
                print(str(excel_data['dob'].strftime('%m/%d/%Y, %H:%M:%S')), flush=True)
                dob = re.search('(((0)[1-9])|((1)[0-2]))(\/)([0-2][0-9]|(3)[0-1])(\/)\d{4}',
                                str(excel_data['dob'].strftime('%m/%d/%Y, %H:%M:%S')))
                gender = re.search('(Nam|Nữ)', str(excel_data['gender']))
                courseID = re.search('^[K|k][1-9][0-9][A-Za-z]+[1-9]*', excel_data['courseID'])
                # print(ID, flush=True)
                # print(fullname, flush=True)
                # print(dob, flush=True)
                # print(gender, flush=True)
                # print(courseID, flush=True)

                if (ID is not None) and (fullname is not None) and (dob is not None) and (gender is not None) and (
                        courseID is not None):

                    User.create(str(excel_data['ID']).replace(' ', ''),
                                str(excel_data['ID']).replace(' ', '') + '@vnu.edu.vn',
                                str(excel_data['ID']).replace(' ', ''),
                                str(excel_data['fullname']).title(),
                                excel_data['dob'],
                                excel_data['gender'],
                                excel_data['courseID'],
                                'Student')
                else:
                    return jsonify({'status': 'error'}), 400

        # When the excel file are included of only student subject status
        # This requires student account to have been existed before being added
        elif max_column == 4:
            print('OK', flush=True)
            for i in range(2, max_row + 1):
                excel_data = {
                    'ID': None,
                    'subjectID': None,
                    'subjectTitle': None,
                    'status': None,
                }
                # Iterate over the dict and all the columns
                for j, index in zip(range(1, max_column + 1), excel_data):
                    # add data to excel_data dict
                    excel_data[index] = sheet.cell(row=i, column=j).value

                print(excel_data, flush=True)
                ID = re.search('^\d{8}$', str(excel_data['ID']).replace(' ', ''))
                if ID is not None:
                    check_student = User.isExist(str(excel_data['ID']).replace(' ', ''))
                    if check_student is True:
                        # add subject to database
                        # check validation
                        subjectID = re.search('(^(([A-Z]|[a-z]){3})([1-9][(0-9)]{3}))',
                                              str(excel_data['subjectID']))
                        status = re.search('(đủ điều kiện|không đủ điều kiện)', str(excel_data['status']).lower())

                        if (subjectID is not None) and (status is not None):
                            Subject.create(excel_data['subjectID'],
                                           excel_data['subjectTitle'])
                            # add qualified and unqualified students to database
                            Student_Status.create(str(excel_data['ID']).replace(' ', ''),
                                                  excel_data['subjectID'],
                                                  excel_data['status'])
                        else:
                            return jsonify({'status': 'bad-request'}), 400
                    else:
                        return jsonify({'status': 'forbidden', 'caution': 'Yêu cầu dữ liệu về thông tin phải tồn tại '
                                                                          'trước rồi mới nhập được danh sách tình trạng '
                                                                          'môn thi của sinh viên'}), 403
                else:
                    return jsonify({'status': 'bad-request'}), 400
        else:
            return jsonify({'status': 'bad-request'}), 400

        Log.create(current_user['ID'],
                   'Tải dữ liệu file Excel lên hệ thống.',
                   set_custom_log_time())

        # when the loop ends, which means all the data has been saved successfully
        return jsonify({'status': 'success'}), 200
    except:
        return jsonify({'status': 'bad-request'}), 400
